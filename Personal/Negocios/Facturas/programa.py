import os
import xml.etree.ElementTree as ET
import openpyxl

# Carpeta que contiene los archivos XML
carpeta = 'xml'

# Obtener la lista de archivos XML en la carpeta
archivos_xml = [archivo for archivo in os.listdir(carpeta) if archivo.endswith('.xml')]

# Crear un nuevo archivo de Excel
libro_excel = openpyxl.Workbook()
hoja = libro_excel.active

# Establecer los encabezados de las columnas
hoja['A1'] = 'Fecha de Emisión'
hoja['B1'] = 'Razón Social'
hoja['C1'] = 'Número Factura'
hoja['D1'] = 'Subtotal 12'
hoja['E1'] = 'Subtotal 0'
hoja['F1'] = 'Valor IVA'
hoja['G1'] = 'TOTAL'
hoja['H1'] = 'Descripción'

# Contador para rastrear la fila actual en la hoja de Excel
fila_actual = 2

# Iterar sobre los archivos XML
for archivo_xml in archivos_xml:
    ruta_xml = os.path.join(carpeta, archivo_xml)
    
    # Parsear el archivo XML
    arbol_xml = ET.parse(ruta_xml)
    raiz = arbol_xml.getroot()
    
    # Extraer la información del XML
    cdata = raiz.find('comprobante').text
    comprobante_root = ET.fromstring(cdata)
    razon_social = comprobante_root.find('.//razonSocial').text
    fecha_emision = comprobante_root.find('.//fechaEmision').text
    cod_Doc = comprobante_root.find('.//codDoc').text
    estab = comprobante_root.find('.//estab').text
    ptoEmi = comprobante_root.find('.//ptoEmi').text
    secuencial = comprobante_root.find('.//secuencial').text
    num_factura = f'{cod_Doc}-{estab}-{ptoEmi}-{secuencial}'
    detalles = comprobante_root.find('.//detalles')

    subtotal0 = 0
    subtotal12 = 0
    ivatotal = 0
    descripcionTotal = ''
    
    # Recorrer los detalles y extraer la información
    for detalle in detalles:
        codigoPrincipal = detalle.find('codigoPrincipal').text
        descripcion = detalle.find('descripcion').text
        cantidad = float(detalle.find('cantidad').text)
        precioUnitario = float(detalle.find('precioUnitario').text)
        descuento = float(detalle.find('descuento').text)
        precioTotalSinImpuesto = detalle.find('precioTotalSinImpuesto').text
        tarifa = float(detalle.find('.//tarifa').text)
        baseImponible = float(detalle.find('.//baseImponible').text)
        valorIVA = float(detalle.find('.//valor').text)
        if descripcionTotal=='':
            descripcionTotal = descripcion
        else:
            descripcionTotal = f'{descripcionTotal}, {descripcion}'

        if tarifa==0:
            subtotal0 = subtotal0 + baseImponible - descuento
        elif tarifa == 12.00:
            subtotal12 = subtotal12 + baseImponible - descuento
            ivatotal = ivatotal + valorIVA
    
    total12 = subtotal12 + ivatotal
    totalfinal = total12 + subtotal0
    
    # Escribir los valores extraídos en la hoja de Excel
    hoja.cell(row=fila_actual, column=1, value=fecha_emision)
    hoja.cell(row=fila_actual, column=2, value=razon_social)
    hoja.cell(row=fila_actual, column=3, value=num_factura)
    hoja.cell(row=fila_actual, column=4, value=subtotal12)
    hoja.cell(row=fila_actual, column=5, value=subtotal0)
    hoja.cell(row=fila_actual, column=6, value=ivatotal)
    hoja.cell(row=fila_actual, column=7, value=totalfinal)
    hoja.cell(row=fila_actual, column=8, value=descripcionTotal)
    
    fila_actual += 1

# Guardar el archivo de Excel
libro_excel.save('datos_factura.xlsx')


